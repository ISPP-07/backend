import os
import re
from typing import Optional
from uuid import uuid4
from datetime import date

from pydantic import UUID4, ValidationError
from fastapi import HTTPException, status, UploadFile
import openpyxl

from src.core.utils.helpers import parse_validation_error, generate_alias, get_valid_mongo_obj
from src.core.deps import DataBaseDep
from src.core.database.base_crud import BulkOperation
from src.modules.acat.patient import model, service
from src.modules.acat.intervention import model as intervention_model, service as intervention_service


async def get_patients_controller(
    db: DataBaseDep,
    alias: Optional[str],
    name: Optional[str],
    nid: Optional[str],
    is_rehabilitated: Optional[bool],
    before_registration_date: Optional[date] = None,
    after_registration_date: Optional[date] = None,
    limit: int = 100,
    offset: int = 0
) -> model.GetPatients:
    patients = await service.get_patients_service(
        db,
        None,
        (
            'alias', {
                '$regex': re.compile(f'^{alias}', re.IGNORECASE)
            } if alias is not None else alias
        ),
        (
            'name', {
                '$regex': re.compile(f'^{name}', re.IGNORECASE)
            } if name is not None else name
        ),
        ('nid', nid),
        ('is_rehabilitated', is_rehabilitated),
        (
            'registration_date', {
                '$lte': before_registration_date.isoformat()
            } if before_registration_date is not None else None
        ),
        (
            'registration_date', {
                '$gte': after_registration_date.isoformat()
            } if after_registration_date is not None else None
        ),
        limit=limit,
        skip=offset
    )
    return model.GetPatients(
        elements=patients,
        total_elements=await service.count_patients_service(db, query={})
    )


async def create_patient_controller(db: DataBaseDep, patient: model.PatientCreate) -> model.Patient:
    check_patient = await service.get_patient_service(db, query={'nid': patient.nid})
    if check_patient is not None:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f'There is already one patient with nid {patient.nid}'
        )
    if patient.alias is None:
        patient.alias = generate_alias(
            patient.name, patient.first_surname, patient.second_surname
        )
    mongo_insert = await service.create_patient_service(db, patient)
    result = await service.get_patient_service(db, query={'id': mongo_insert.inserted_id})
    return result


async def get_patient_details_controller(db: DataBaseDep, patient_id: UUID4):
    result = await service.get_patient_service(db, query={'id': patient_id})
    if result is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail='Patient not found',
        )
    return result


async def upload_excel_patients_controller(db: DataBaseDep, patients: UploadFile) -> None:
    [_, extension] = os.path.splitext(patients.filename)
    if extension[1:] not in ['xlsx', 'xlsm']:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                'The files with extension ',
                f'"{extension[1:]}" are not supported.'
            )
        )
    fields_excel = [
        'nombre',
        'primer apellido',
        'segundo apellido',
        'dni',
        'fecha nacimiento',
        'genero',
        'direccion',
        'telefono',
        'numero expediente',
        'tecnico',
        'observacion']
    wb = openpyxl.load_workbook(patients.file)
    ws = wb.active
    first_row = [
        ws.cell(row=1, column=i).value
        for i in range(1, len(fields_excel) + 1)
    ]
    if len(first_row) != len(fields_excel) and not all(
            field in fields_excel for field in first_row):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail='The excel file is incorrect'
        )
    patients_excel: list[model.PatientCreate] = []
    for row in ws.iter_rows(
            min_row=2,
            min_col=1,
            max_col=11,
            values_only=True):
        if all(value is None for value in row):
            continue
        if row[0] is None or row[1] is None or row[3] is None or row[4] is None or row[8] is None:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail='The excel file is incorrect'
            )
        if row[5] is not None and row[5] not in ['Hombre', 'Mujer']:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail='The excel file is incorrect'
            )
        try:
            new_patient = model.PatientCreate(
                name=row[0],
                first_surname=row[1],
                second_surname=row[2],
                alias=generate_alias(row[0], row[1], row[2]),
                nid=row[3],
                birth_date=row[4],
                gender=(
                    'Man' if row[5] == 'Hombre' else 'Woman'
                ) if row[5] is not None else None,
                address=row[6],
                contact_phone=str(row[7]),
                dossier_number=row[8],
                first_technician=row[9],
                observation=row[10],
            )
        except ValidationError as e:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=parse_validation_error(e.errors())
            )
        patients_excel.append(new_patient)
    patients_db = await service.get_patients_service(db, query={'nid': {'$in': [p.nid for p in patients_excel]}})
    nids_db = [p.nid for p in patients_db]
    patients_create = [
        BulkOperation(
            bulk_type='InsertOne',
            data=model.Patient(**p.model_dump(), id=uuid4()).mongo()
        )
        for p in patients_excel
        if p.nid not in nids_db
    ]
    patients_update = [
        BulkOperation(
            bulk_type='UpdateOne',
            data={'$set': get_valid_mongo_obj(p.model_dump())},
            query=model.Patient.prepare_query({'nid': p.nid})
        )
        for p in patients_excel
        if p.nid in nids_db
    ]
    interventions_update = [
        BulkOperation(
            bulk_type='UpdateMany',
            data={'$set': {'patient': p.data['$set']}},
            query=intervention_model.Intervention.prepare_query(
                {'patient.nid': p.data['$set']['nid']}
            )
        )
        for p in patients_update
    ]
    patients_operations = patients_create + patients_update
    if len(patients_operations) > 0:
        await service.bulk_service(db, operations=patients_operations, ordered=False)
    if len(interventions_update) > 0:
        await intervention_service.bulk_service(db, operations=interventions_update, ordered=False)


async def update_patient_controller(
    db: DataBaseDep,
    patient_id: UUID4,
    patient: model.PatientUpdate
) -> model.Patient:
    patient_db = await service.get_patient_service(db, query={'id': patient_id})
    if patient_db is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail='Patient not found',
        )
    if patient.nid is not None:
        check_nid = await service.get_patient_service(db, query={'nid': patient.nid})
        if check_nid is not None and check_nid.id != patient_id:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f'There is already a patient with nid {patient.nid}',
            )
    request_none_fields = [
        field for field in model.PATIENT_NONE_FIELDS
        if field in patient.update_fields_to_none
    ]
    update_data = patient.model_dump(exclude='update_fields_to_none')
    for field in update_data.copy():
        if field in request_none_fields:
            continue
        if update_data[field] is None:
            update_data.pop(field)
    updated_patient = await service.update_patient_service(db, {'id': patient_id}, update_data)
    if updated_patient is not None:
        await intervention_service.update_interventions_service(
            db,
            {'patient.id': patient_id},
            {'patient': updated_patient.model_dump()}
        )
    return updated_patient


async def delete_patient_controller(db: DataBaseDep, patient_id: UUID4):
    await service.delete_patient_service(db, query={'id': patient_id})
