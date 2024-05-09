import re
import os
from typing import Optional
from datetime import date
from uuid import uuid4

import openpyxl
from pydantic import UUID4, ValidationError
from fastapi import HTTPException, status, UploadFile

from src.core.deps import DataBaseDep
from src.core.database.base_crud import BulkOperation
from src.core.utils.helpers import parse_validation_error
from src.modules.acat.intervention import service, model
from src.modules.acat.patient import service as patient_service, model as patient_model


async def get_interventions_controller(
    db: DataBaseDep,
    before_date: Optional[date],
    after_date: Optional[date],
    technician: Optional[str],
    patient: Optional[UUID4],
    limit: int = 100,
    offset: int = 0
) -> model.GetInterventions:
    interventions = await service.get_interventions_service(
        db,
        (
            'date', {
                '$lte': before_date.isoformat()
            } if before_date is not None else None
        ),
        (
            'date', {
                '$gte': after_date.isoformat()
            } if after_date is not None else None
        ),
        (
            'technician', {
                '$regex': re.compile(f'^{technician}', re.IGNORECASE)
            } if technician is not None else technician
        ),
        (
            'patient.id', patient
        ),
        limit=limit,
        skip=offset
    )
    return model.GetInterventions(
        elements=interventions,
        total_elements=await service.count_interventions_service(db, query={})
    )


async def get_intervention_details_controller(db: DataBaseDep, intervention_id: UUID4):
    result = await service.get_intervention_service(db, query={'id': intervention_id})
    if result is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Intervention not found",
        )
    return result


async def create_intervention_controller(db: DataBaseDep, intervention: model.InterventionCreate):

    patient = await patient_service.get_patient_service(db, query={'id': intervention.patient_id})

    if patient is None:
        raise HTTPException(status_code=404, detail="Patient not found")

    mongo_insert = await service.create_intervention_service(db, intervention, patient)
    result = await service.get_intervention_service(db, query={'id': mongo_insert.inserted_id})

    return result


async def update_intervention_controller(
    db: DataBaseDep,
    intervention_id: UUID4,
    intervention: model.InterventionUpdate
) -> model.Intervention:
    intervention_db = await service.get_intervention_service(db, query={'id': intervention_id})
    if intervention_db is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail='Intervention not found',
        )
    request_none_fields = [
        field for field in model.INTERVENTION_NONE_FIELDS
        if field in intervention.update_fields_to_none
    ]
    update_data = intervention.model_dump(
        exclude={'update_fields_to_none', 'patient_id'}
    )
    if intervention.patient_id is not None:
        patient_db = await patient_service.get_patient_by_id(db, intervention.patient_id)
        if patient_db is None:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail='Patient not found',
            )
        update_data['patient'] = patient_db.model_dump()
    for field in update_data.copy():
        if field in request_none_fields:
            update_data[field] = None
            continue
        if update_data[field] is None:
            update_data.pop(field)
    updated_intervention = await service.update_intervention_service(
        db,
        intervention_id,
        update_data
    )
    return updated_intervention


async def delete_intervention_controller(db: DataBaseDep, intervention_id: UUID4):
    await service.delete_intervention_service(db, query={'id': intervention_id})


async def upload_excel_interventions_controller(db: DataBaseDep, patients: UploadFile) -> None:
    [_, extension] = os.path.splitext(patients.filename)
    if extension[1:] not in ['xlsx', 'xlsm']:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                'The files with extension ',
                f'"{extension[1:]}" are not supported.'
            )
        )
    fields_excel = [
        'fecha',
        'motivo',
        'tipo',
        'observacion',
        'tecnico',
        'dni beneficiario'
    ]
    wb = openpyxl.load_workbook(patients.file)
    ws = wb.active
    first_row = [
        ws.cell(row=1, column=i).value
        for i in range(1, len(fields_excel) + 1)
    ]
    if not all(
        field in fields_excel for field in first_row
    ):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail='The excel file is incorrect'
        )
    interventions_excel: list[model.Intervention] = []
    for row in ws.iter_rows(
        min_row=2,
        min_col=1,
        max_col=6,
        values_only=True
    ):
        if all(value is None for value in row):
            continue
        if row[0] is None or row[4] is None or row[5] is None:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail='The excel file is incorrect'
            )
        patient = await patient_service.get_patient_service(db, {'nid': row[5]})
        if patient is None:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f'Patient with nid {row[5]} not found'
            )
        try:
            new_intervention = model.Intervention(
                id=uuid4(),
                date=row[0],
                reason=row[1],
                typology=row[2],
                observations=row[3],
                technician=row[4],
                patient=patient_model.Patient(**patient.model_dump())
            )
        except ValidationError as e:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=parse_validation_error(e.errors())
            )
        interventions_excel.append(new_intervention)
    intervention_operations = [
        BulkOperation(
            bulk_type='InsertOne',
            data=intervention.mongo()
        )
        for intervention in interventions_excel
    ]
    if len(intervention_operations) > 0:
        await service.bulk_service(db, operations=intervention_operations, ordered=False)
