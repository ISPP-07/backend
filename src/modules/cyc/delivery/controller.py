import os
from collections import Counter
from uuid import uuid4
from typing import Dict, Optional
from datetime import date

import openpyxl
from pydantic import UUID4, ValidationError
from fastapi import HTTPException, status, UploadFile

from src.core.deps import DataBaseDep
from src.core.database.base_crud import BulkOperation
from src.core.utils.helpers import parse_validation_error, get_valid_mongo_obj
from src.modules.cyc.delivery import model, service
from src.modules.cyc.family import service as family_service
from src.modules.cyc.warehouse import service as product_service, model as product_model


async def get_deliveries_controller(
    db: DataBaseDep,
    before_date: Optional[date],
    after_date: Optional[date],
    state: Optional[model.State],
    family: Optional[UUID4],
    limit: int = 100,
    offset: int = 0
) -> model.GetDeliveries:
    deliveries = await service.get_deliveries_service(
        db,
        (
            'date', {
                '$lte': before_date.isoformat()
            } if before_date is not None else None
        ),
        (
            'date', {
                '$gte': after_date.isoformat()
            } if after_date is not None else None
        ),
        (
            'state', state.value if state is not None else state
        ),
        (
            'family_id', family
        ),
        limit=limit,
        skip=offset
    )
    warehouses = await product_service.get_warehouses_service(db, query=None)
    product_to_name = {
        product.id: product.name for warehouse in warehouses for product in warehouse.products
    }
    deliveries_out = []
    for delivery in deliveries:
        updated_lines = []
        for line in delivery.lines:
            product_name = product_to_name.get(line.product_id)
            updated_line = model.DeliveryLineOut(
                **line.model_dump(), name=product_name
            )
            updated_lines.append(updated_line)
        out = model.DeliveryOut(
            id=delivery.id,
            date=delivery.date,
            months=delivery.months,
            state=delivery.state,
            lines=updated_lines,
            family_id=delivery.family_id
        )
        deliveries_out.append(out)
    return model.GetDeliveries(
        elements=deliveries_out,
        total_elements=await service.count_deliveries_service(db, query={})
    )


async def get_delivery_details_controller(db: DataBaseDep, delivery_id: int) -> model.DeliveryOut:
    result = await service.get_delivery_service(db, query={'id': delivery_id})
    if result is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail='Delivery not found',
        )
    warehouses = await product_service.get_warehouses_service(db, query=None)
    product_to_name = {
        product.id: product.name for warehouse in warehouses for product in warehouse.products}

    updated_lines = []
    for line in result.lines:
        product_name = product_to_name.get(line.product_id)
        updated_line = model.DeliveryLineOut(
            **line.model_dump(), name=product_name)
        updated_lines.append(updated_line)
    salida = model.DeliveryOut(id=result.id,
                               date=result.date,
                               months=result.months,
                               state=result.state,
                               lines=updated_lines,
                               family_id=result.family_id)
    return salida


async def create_delivery_controller(db: DataBaseDep, create_delivery: model.DeliveryCreate) -> model.Delivery:
    # Validar la unicidad del producto en las líneas
    products_count = Counter(line.product_id for line in create_delivery.lines)
    if any(count > 1 for count in products_count.values()):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail='There is one product that is in two different lines. ' +
            'Please put them in a single line'
        )

    # Asegurar que la familia existe
    family = await family_service.get_family_service(db, query={'id': create_delivery.family_id})
    if family is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f'Family {create_delivery.family_id} not found'
        )

    # Recuperar almacenes y validar que todos los productos existan y tengan
    # suficiente stock
    warehouses = await product_service.get_warehouses_service(db, query=None)
    product_to_warehouse: Dict[str, tuple] = {product.id: (
        warehouse, product) for warehouse in warehouses for product in warehouse.products}

    missing_products = [
        product_id for product_id in products_count if product_id not in product_to_warehouse]
    if missing_products:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail='Product not found in any warehouse'
        )

    for product_id, required_quantity in (
            (line.product_id, line.quantity) for line in create_delivery.lines):
        warehouse, product = product_to_warehouse.get(product_id, (None, None))
        if not warehouse or product.quantity < required_quantity:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f'Not enough stock for product {product.name}. ' +
                f'There is only {product.quantity} left' if warehouse else 'Product not found')

    # Preparar las actualizaciones por producto
    product_updates = {}  # Diccionario para acumular actualizaciones
    for line in create_delivery.lines:
        warehouse, product = product_to_warehouse[line.product_id]
        if product.id not in product_updates:
            product_updates[product.id] = product.quantity - line.quantity
        else:
            product_updates[product.id] -= line.quantity
    for warehouse in warehouses:
        updated_products = []
        for product in warehouse.products:
            if product.id in product_updates:
                updated_product = product_model.Product(
                    id=product.id,
                    name=product.name,
                    quantity=product_updates[product.id],
                    exp_date=product.exp_date
                )
            else:
                updated_product = product
            updated_products.append(updated_product.model_dump())

        await product_service.update_warehouse_service(
            db,
            warehouse_id=warehouse.id,
            warehouse_update={'products': updated_products}
        )
    mongo_insert = await service.create_delivery_service(db, create_delivery)
    result = await service.get_delivery_service(db, query={'id': mongo_insert.inserted_id})
    return result


async def update_delivery_controller(db: DataBaseDep, delivery_id: UUID4, delivery: model.DeliveryUpdate) -> model.Delivery:
    delivery_actual = await service.get_delivery_service(db, query={'id': delivery_id})
    if delivery_actual is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail='Delivery not found'
        )

    if delivery_actual.state == model.State.DELIVERED and delivery.state is None:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail='Cannot update a delivered delivery'
        )

    # Validar la unicidad del producto en las líneas
    if delivery.lines:
        products_count = Counter(line.product_id for line in delivery.lines)
        if any(count > 1 for count in products_count.values()):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail='There is one product that is in two different lines. ' +
                'Please put them in a single line')

        warehouses = await product_service.get_warehouses_service(db, query=None)
        product_to_warehouse = {
            product.id: (
                warehouse,
                product) for warehouse in warehouses for product in warehouse.products}

        missing_products = [
            product_id for product_id in products_count if product_id not in product_to_warehouse]
        if missing_products:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail='Product not found in any warehouse'
            )

        product_updates = {}  # Diccionario para acumular diferencias de cantidad
        for line in delivery.lines:
            warehouse, product = product_to_warehouse.get(
                line.product_id, (None, None))
            if not warehouse or product is None:
                continue  # Omitir si el producto no se encuentra en ningún almacén

            # Calcula la diferencia de cantidad basada en la entrega actual y la
            # nueva
            old_quantity = next(
                (old_line.quantity for old_line in delivery_actual.lines if old_line.product_id == line.product_id),
                0)
            quantity_difference = old_quantity - line.quantity

            if product.id not in product_updates:
                product_updates[product.id] = product.quantity + \
                    quantity_difference
            else:
                product_updates[product.id] += quantity_difference

            if product_updates[product.id] < 0:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f'Not enough stock for product {product.name}.' +
                    f' There is only {product.quantity} left')

        # Proceder con la actualización de las cantidades de los productos en sus
        # respectivos almacenes
        for warehouse in warehouses:
            updated_products = []
            for product in warehouse.products:
                if product.id in product_updates:
                    updated_product = product_model.Product(
                        id=product.id,
                        name=product.name,
                        quantity=product_updates[product.id],
                        exp_date=product.exp_date
                    )
                else:
                    updated_product = product
                updated_products.append(updated_product.model_dump())

            await product_service.update_warehouse_service(
                db,
                warehouse_id=warehouse.id,
                warehouse_update={'products': updated_products}
            )

    # Asegurar que la familia existe
    if delivery.family_id:
        family = await family_service.get_family_service(db, query={'id': delivery.family_id})
        if family is None:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f'Family {delivery.family_id} not found'
            )

    update_data = delivery.model_dump()
    for field in update_data.copy():
        if update_data[field] is None:
            update_data.pop(field)
    result = await service.update_delivery_service(db, query={'id': delivery_id}, delivery=update_data)
    return result


async def delete_delivery_controller(db: DataBaseDep, delivery_id: UUID4):
    return await service.delete_delivery_service(db, query={'id': delivery_id})


async def get_family_deliveries_controller(db: DataBaseDep, family_id: int) -> list[model.DeliveryOut]:
    deliveries = await service.get_deliveries_service(db)
    result = [
        delivery for delivery in deliveries if delivery.family_id == family_id]
    warehouses = await product_service.get_warehouses_service(db, query=None)
    product_to_name = {
        product.id: product.name for warehouse in warehouses for product in warehouse.products}
    result_final = []
    for delivery in result:
        updated_lines = []
        for line in delivery.lines:
            product_name = product_to_name.get(line.product_id)
            updated_line = model.DeliveryLineOut(
                **line.model_dump(), name=product_name)
            updated_lines.append(updated_line)
        salida = model.DeliveryOut(id=delivery.id,
                                   date=delivery.date,
                                   months=delivery.months,
                                   state=delivery.state,
                                   lines=updated_lines,
                                   family_id=delivery.family_id)
        result_final.append(salida)
    return result_final


async def upload_excel_deliveries_controller(db: DataBaseDep, deliveries: UploadFile) -> None:
    [_, extension] = os.path.splitext(deliveries.filename)
    if extension[1:] not in ['xlsx', 'xlsm']:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                'The files with extension ',
                f'"{extension[1:]}" are not supported.'
            )
        )
    fields_delivery_excel = [
        'numero entrega', 'fecha', 'meses',
        'estado', 'documento identidad cabeza familia'
    ]
    fields_line_excel = [
        'numero entrega', 'almacen producto',
        'nombre producto', 'cantidad', 'estado'
    ]
    wb = openpyxl.load_workbook(deliveries.file)
    ws = wb.active
    first_row_delivery = [
        ws.cell(row=1, column=i).value
        for i in range(1, len(fields_delivery_excel) + 1)
    ]
    first_row_line = [
        ws.cell(row=1, column=i).value
        for i in range(7, 6 + len(fields_line_excel) + 1)
    ]
    if not all(
        field in fields_delivery_excel for field in first_row_delivery
    ) and not all(
        field in fields_line_excel for field in first_row_line
    ):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail='The excel file is incorrect'
        )
    lines_excel: dict[int, list[model.DeliveryLine]] = {}
    deliveries_excel: list[model.Delivery] = []
    warehouses = await product_service.get_warehouses_service(db)
    updated_products: dict[UUID4, list[product_model.Product]] = {}
    for row in ws.iter_rows(
        min_row=2,
        min_col=7,
        max_col=11,
        values_only=True
    ):
        if all(value is None for value in row):
            continue
        if row[0] is None or row[1] is None or row[2] is None or row[3] is None or not isinstance(row[3], int):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail='The excel file is incorrect'
            )
        warehouse = next(
            (w for w in warehouses if w.name == str(row[1])),
            None
        )
        if warehouse is None:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f'Warehouse with name {row[1]} not found'
            )
        product = next(
            (p for p in warehouse.products if p.name == row[2]),
            None
        )
        if product is None:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f'Product with name {row[2]} not found'
            )
        if row[0] in lines_excel and product.id in [l.product_id for l in lines_excel[row[0]]]:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f'Product {product.name} is twice in the delivery'
            )
        if product.quantity - row[3] < 0:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=(
                    "There aren't enough products"
                    f"{product.name} for the delivery"
                )
            )
        updated_product = product_model.Product(
            **product.model_dump(exclude=['quantity']),
            quantity=product.quantity - row[3]
        )
        if (warehouse.id not in updated_products):
            products = warehouse.products.copy()
            products.pop(products.index(product))
            updated_products[warehouse.id] = products + [updated_product]
        else:
            updated_products[warehouse.id].pop(
                updated_products[warehouse.id].index(product)
            )
            updated_products[warehouse.id].append(updated_product)
        try:
            new_line = model.DeliveryLine(
                product_id=product.id,
                quantity=row[3],
                state=row[4]
            )
        except ValidationError as e:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=parse_validation_error(e.errors())
            )
        if (row[0] not in lines_excel):
            lines_excel[row[0]] = [new_line]
        else:
            lines_excel[row[0]].append(new_line)
    families = await family_service.get_families_service(db)
    for row in ws.iter_rows(
        min_row=2,
        min_col=1,
        max_col=5,
        values_only=True
    ):
        if all(value is None for value in row):
            continue
        if row[0] is None or row[1] is None or row[2] is None or row[3] is None or row[4] is None:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail='The excel file is incorrect'
            )
        state_value = None
        if row[3] not in ['Espera', 'Notificado', 'Entregado']:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail='The excel file is incorrect'
            )
        else:
            if row[3] == 'Espera':
                state_value = model.State.NEXT
            elif row[3] == 'Notificado':
                state_value = model.State.NOTIFIED
            else:
                state_value = model.State.DELIVERED
        family = next(
            (f for f in families if row[4] in [m.nid for m in f.members]),
            None
        )
        if family is None:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=(
                    "Family with family's head"
                    f"identifer {row[4]} not found"
                )
            )
        try:
            new_delivery = model.Delivery(
                id=uuid4(),
                date=row[1],
                months=row[2],
                state=state_value,
                lines=lines_excel[row[0]],
                family_id=family.id
            )
        except ValidationError as e:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=parse_validation_error(e.errors())
            )
        deliveries_excel.append(new_delivery)
    warehouse_operations = [
        BulkOperation(
            bulk_type='UpdateOne',
            data={'$set': {
                'products': get_valid_mongo_obj([p.model_dump() for p in u_products])
            }},
            query=product_model.Warehouse.prepare_query(
                {'id': w_id}
            )
        )
        for w_id, u_products in updated_products.items()
    ]
    deliveries_operations = [
        BulkOperation(
            bulk_type='InsertOne',
            data=d.mongo()
        )
        for d in deliveries_excel
    ]
    if len(deliveries_operations) > 0:
        await service.bulk_service(db, operations=deliveries_operations, ordered=False)
    if len(warehouse_operations) > 0:
        await product_service.bulk_service(db, operations=warehouse_operations, ordered=False)
