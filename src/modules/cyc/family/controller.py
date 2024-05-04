import re
import os
from uuid import uuid4
from typing import Optional

import openpyxl
from fastapi import HTTPException, status, UploadFile
from pydantic import UUID4, ValidationError

from src.core.deps import DataBaseDep
from src.core.database.base_crud import BulkOperation
from src.core.utils.helpers import parse_validation_error
from src.modules.cyc.family import model
from src.modules.cyc.family import service


async def get_families_controller(
    db: DataBaseDep,
    state: Optional[model.DerecognitionStatus],
    referred_organization: Optional[str],
    name: Optional[str],
    limit=100,
    offset=0
) -> model.GetFamilies:
    families = await service.get_families_service(
        db,
        ('derecognition_state', state.value if state is not None else state),
        (
            'referred_organization', {
                '$regex': re.compile(f'^{referred_organization}', re.IGNORECASE)
            } if referred_organization is not None else referred_organization
        ),
        (
            'name', {
                '$regex': re.compile(f'^{name}', re.IGNORECASE)
            } if name is not None else name
        ),
        limit=limit,
        skip=offset
    )
    return model.GetFamilies(
        elements=families,
        total_elements=await service.count_families_service(db, query={})
    )


async def create_family_controller(db: DataBaseDep, family: model.FamilyCreate) -> model.Family:
    if family.members is not None:
        persons = await service.get_members_service(db)
        for m in family.members:
            if m.nid is not None and any(
                    person.nid == m.nid for person in persons):
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f'There is already a person with this NID: {m.nid}',
                )
    # TENGO QUE ELIMINAR PASSPORT FIELD DE PERSON
    new_family = family.model_dump()
    for member in new_family['members']:
        del member['passport']
    mongo_insert = await service.create_family_service(db, new_family)
    result = await service.get_family_service(db, query={'id': mongo_insert.inserted_id})
    return result


async def get_family_details_controller(db: DataBaseDep, family_id: int) -> model.Family:
    result = await service.get_family_service(db, query={'id': family_id})
    if result is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail='Family not found',
        )
    return result


async def update_family_controller(db: DataBaseDep, family_id: UUID4, family: model.FamilyUpdate) -> model.Family:
    # if family.members is not None:
    #     persons = await service.get_members_service(db)
    #     for m in family.members:
    #         if m.nid is not None and any(
    #                 person.nid == m.nid for person in persons):
    #             raise HTTPException(
    #                 status_code=status.HTTP_400_BAD_REQUEST,
    #                 detail=f'There is already a person with this NID: {m.nid}',
    #             )
    request_none_fields = [
        field for field in model.FAMILY_NONE_FIELDS
        if field in family.update_fields_to_none
    ]
    update_data = family.model_dump(
        exclude=['update_fields_to_none']
    )
    for field in update_data.copy():
        if field in request_none_fields:
            continue
        if field == 'members' and update_data[field] is not None:
            for member in update_data[field]:
                del member['passport']
        if update_data[field] is None:
            update_data.pop(field)
    result = await service.update_family_service(db, family_id=family_id, family_update=update_data)
    if result is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail='Family not found',
        )
    return result


async def delete_family_controller(db: DataBaseDep, family_id: UUID4):
    await service.delete_family_service(db, query={'id': family_id})


async def update_person_controller(db: DataBaseDep, family_id: UUID4, person_nid: str, person: model.PersonUpdate) -> model.Person:
    family = await service.get_family_service(db, query={'id': family_id})
    if family is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail='Family not found',
        )
    old_person = next(
        (person for person in family.members if person.nid == person_nid),
        None
    )
    if old_person is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail='Person not found in family',
        )
    request_none_fields = [
        field for field in model.PERSON_NONE_FIELDS
        if field in person.update_fields_to_none
    ]
    update_data = person.model_dump(
        exclude=['passport', 'update_fields_to_none']
    )
    old_person_data = old_person.model_dump()
    for field in old_person_data:
        if field in request_none_fields:
            continue
        if field not in update_data or update_data[field] is None:
            update_data[field] = old_person_data[field]
    members = [
        p.model_dump() for p in family.members
        if p.nid != old_person.nid
    ]
    members.append(update_data)
    return await service.update_family_service(
        db,
        family_id=family.id,
        family_update={'members': members}
    )


async def delete_person_controller(db: DataBaseDep, family_id: UUID4, person_nid: str) -> None:
    new_family = await service.get_family_service(db, query={'id': family_id})
    if new_family is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail='Family not found',
        )

    if person_nid not in [person.nid for person in new_family.members]:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail='Person not found in family',
        )

    head_person = [
        person for person in new_family.members if person.family_head
    ][0]
    if head_person.nid == person_nid:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail='Cannot delete the family head',
        )
    members = [
        person.model_dump() for person in new_family.members
        if person.nid != person_nid
    ]
    await service.update_family_service(
        db,
        family_id,
        family_update={'members': members}
    )


async def upload_excel_families_controller(db: DataBaseDep, families: UploadFile) -> None:
    [_, extension] = os.path.splitext(families.filename)
    if extension[1:] not in ['xlsx', 'xlsm']:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                'The files with extension ',
                f'"{extension[1:]}" are not supported.'
            )
        )
    fields_familie_excel = [
        'numero familia', 'nombre', 'numero telefono', 'direccion',
        'fecha renovacion', 'estado', 'organizacion referida', 'observacion'
    ]
    fields_person_excel = [
        'numero familia',
        'fecha nacimiento',
        'nombre',
        'apellido',
        'nacionalidad',
        'documento identidad',
        'cabeza familia',
        'genero',
        'diversidad funcional',
        'intolerancia alimenticia',
        'sin hogar']
    wb = openpyxl.load_workbook(families.file)
    ws = wb.active
    first_row_familie = [
        ws.cell(row=1, column=i).value
        for i in range(1, len(fields_familie_excel) + 1)
    ]
    first_row_person = [
        ws.cell(row=1, column=i).value
        for i in range(10, 9 + len(fields_person_excel) + 1)
    ]
    if not all(
        field in fields_familie_excel for field in first_row_familie
    ) and not all(
        field in fields_person_excel for field in first_row_person
    ):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail='The excel file is incorrect'
        )
    persons_excel: dict[int, list[model.PersonCreate]] = {}
    families_excel: list[model.Family] = []
    for row in ws.iter_rows(
        min_row=2,
        min_col=10,
        max_col=20,
        values_only=True
    ):
        if all(value is None for value in row):
            continue
        if row[0] is None or row[1] is None or row[4] is None or row[7] is None:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail='The excel file is incorrect'
            )
        if row[7] not in ['Hombre', 'Mujer']:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail='The excel file is incorrect'
            )
        parsed_nid: str | None = row[5]
        is_passport = False
        if parsed_nid is not None and parsed_nid.startswith('P-'):
            parsed_nid = parsed_nid[2:]
            is_passport = True
        is_family_head = False
        if row[6] is not None:
            is_family_head = True
        has_functional_diversity = False
        if row[8] is not None:
            has_functional_diversity = True
        food_intolerances: list[str] = []
        if row[9] is not None:
            intolerances: str = row[9]
            food_intolerances = intolerances.split(',')
        is_homeless = False
        if row[10] is not None:
            is_homeless = True
        try:
            new_person = model.PersonCreate(
                date_birth=row[1],
                name=row[2],
                surname=row[3],
                nationality=row[4],
                nid=parsed_nid,
                family_head=is_family_head,
                gender='Man' if row[7] == 'Hombre' else 'Woman',
                functional_diversity=has_functional_diversity,
                food_intolerances=food_intolerances,
                homeless=is_homeless,
                passport=is_passport
            )
        except ValidationError as e:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=parse_validation_error(e.errors())
            )
        if (row[0] not in persons_excel):
            persons_excel[row[0]] = [new_person]
        else:
            persons_excel[row[0]].append(new_person)
    for row in ws.iter_rows(
        min_row=2,
        min_col=1,
        max_col=8,
        values_only=True
    ):
        if all(value is None for value in row):
            continue
        if row[0] is None or row[1] is None or row[2] is None or row[3] is None or row[5] is None:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail='The excel file is incorrect'
            )
        state_value = None
        if row[5] not in ['Activa', 'Suspendida']:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail='The excel file is incorrect'
            )
        else:
            if row[5] == 'Activa':
                state_value = model.DerecognitionStatus.ACTIVE
            else:
                state_value = model.DerecognitionStatus.SUSPENDED
        try:
            new_family = model.Family(
                id=uuid4(),
                name=row[1],
                phone=str(row[2]),
                address=row[3],
                next_renewal_date=row[4],
                derecognition_state=state_value,
                referred_organization=row[6],
                observation=row[7],
                members=[
                    model.Person(**p.model_dump(exclude=['passport']))
                    for p in persons_excel[row[0]]
                ]
            )
        except ValidationError as e:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=parse_validation_error(e.errors())
            )
        families_excel.append(new_family)
    families_operations = [
        BulkOperation(
            bulk_type='InsertOne',
            data=f.mongo()
        )
        for f in families_excel
    ]
    if len(families_operations) > 0:
        await service.bulk_service(db, operations=families_operations, ordered=False)
