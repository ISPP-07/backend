import os
from uuid import uuid4
from collections import Counter

from fastapi import HTTPException, status, UploadFile
import openpyxl
from pydantic import UUID4, ValidationError

from src.core.deps import DataBaseDep
from src.core.database.base_crud import BulkOperation
from src.core.utils.helpers import parse_validation_error, get_valid_mongo_obj
from src.modules.cyc.warehouse import service
from src.modules.cyc.warehouse import model


async def get_products_controller(db: DataBaseDep) -> list[model.ProductOut]:
    return await service.get_products_service(db)


async def get_product_controller(db: DataBaseDep, product_id: UUID4) -> model.ProductOut:
    products = await service.get_products_service(db)
    result = next(
        (p for p in products if p.id == product_id),
        None
    )
    if result is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f'Product {product_id} not found'
        )
    return result


async def get_warehouses_controller(db: DataBaseDep) -> list[model.Warehouse]:
    return await service.get_warehouses_service(db)


async def get_warehouse_controller(db: DataBaseDep, warehouse_id: UUID4) -> model.Warehouse:
    return await service.get_warehouse_service(db, query={'id': warehouse_id})


async def create_product_controller(
    db: DataBaseDep,
    create_products: model.ProductCreate
) -> list[model.ProductOut]:
    result = []
    warehouses_id = [p.warehouse_id for p in create_products.products]
    warehouses = await service.get_warehouses_service(db, query={'id': {'$in': warehouses_id}})
    products_by_warehouse: dict[UUID4, list[dict]] = {}
    products_count = Counter([p.name for p in create_products.products])
    if any(count > 1 for count in products_count.values()):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail='There can not be duplicated products'
        )
    for product in create_products.products:
        warehouse = next(
            (w for w in warehouses if w.id if product.warehouse_id),
            None
        )
        if warehouse is None:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f'Warehouse {product.warehouse_id} not found'
            )
        if any(p.name == product.name for p in warehouse.products):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=(f'Product with name {product.name} '
                        f'already exists in warehouse {warehouse.name}')
            )
        new_product = model.Product(
            id=uuid4(),
            name=product.name,
            quantity=product.quantity,
            exp_date=product.exp_date
        ).model_dump()
        if warehouse.id not in products_by_warehouse:
            products_by_warehouse[warehouse.id] = []
        products_by_warehouse[warehouse.id].append(new_product)
        result.append(model.ProductOut(
            id=new_product['id'],
            warehouse_id=warehouse.id,
            name=product.name,
            quantity=product.quantity,
            exp_date=product.exp_date
        ))
    for key, value in products_by_warehouse.items():
        warehouse = [w for w in warehouses if w.id == key][0]
        warehouse_products = [
            p.model_dump() for p
            in warehouse.products
        ]
        await service.update_warehouse_service(
            db,
            warehouse_id=warehouse.id,
            warehouse_update={'products': warehouse_products + value}
        )
    return result


async def create_warehouse_controller(
    db: DataBaseDep,
    warehouse: model.WarehouseCreate
) -> model.Warehouse:
    check_warehouse = await service.get_warehouse_service(db, query={'name': warehouse.name})
    if check_warehouse is not None:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail='Warehouse already created'
        )
    new_warehouse = warehouse.model_dump()
    products_with_id = [
        model.Product(
            id=uuid4(),
            name=p.name,
            quantity=p.quantity,
            exp_date=p.exp_date
        ) for p in warehouse.products
    ]
    new_warehouse['products'] = products_with_id
    mongo_insert = await service.create_warehouse_service(db, new_warehouse)
    result = await service.get_warehouse_service(db, query={'id': mongo_insert.inserted_id})
    return result


async def update_product_controller(
        db: DataBaseDep,
        update_products: model.ProductUpdate
) -> list[model.ProductOut]:
    result = []
    warehouses_id = [p.warehouse_id for p in update_products.products]
    warehouses = await service.get_warehouses_service(db, query={'id': {'$in': warehouses_id}})
    products_by_warehouse: dict[UUID4, list[dict]] = {}
    product_ids_count = Counter(p.product_id for p in update_products.products)
    if any(value > 1 for value in product_ids_count.values()):
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail='There cannot be duplicated products'
        )
    for product in update_products.products:
        warehouse = next(
            (w for w in warehouses if w.id if product.warehouse_id),
            None
        )
        if warehouse is None:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f'Warehouse {product.warehouse_id} not found'
            )
        if product.product_id not in (p.id for p in warehouse.products):
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f'Product not found in warehouse {warehouse.name}'
            )
        old_product = list(filter(
            lambda p: p.id == product.product_id,
            warehouse.products
        )).pop()
        updated_product = model.Product(
            id=product.product_id,
            name=old_product.name if product.name is None else product.name,
            quantity=old_product.quantity if product.quantity is None else product. quantity,
            exp_date=product.exp_date if product.update_exp_date else old_product.exp_date)
        if warehouse.id not in products_by_warehouse:
            products_by_warehouse[warehouse.id] = []
        products_by_warehouse[warehouse.id].append(
            updated_product.model_dump())
        result.append(model.ProductOut(
            id=product.product_id,
            warehouse_id=warehouse.id,
            name=updated_product.name,
            quantity=updated_product.quantity,
            exp_date=updated_product.exp_date
        ))
    for key, value in products_by_warehouse.items():
        warehouse = [w for w in warehouses if w.id == key][0]
        value_ids = [p['id'] for p in value]
        await service.update_warehouse_service(
            db,
            warehouse_id=warehouse.id,
            warehouse_update={
                'products': [p.model_dump() for p in warehouse.products if p.id not in value_ids] + value
            }
        )
    return result


async def delete_warehouse_controller(db: DataBaseDep, warehouse_id: UUID4) -> None:
    await service.delete_warehouse_service(db, warehouse_id)


async def delete_product_controller(db: DataBaseDep, product_id: UUID4) -> None:
    products = await service.get_products_service(db)
    product = next(
        (p for p in products if p.id == product_id),
        None
    )
    if product is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f'Product {product_id} not found'
        )
    warehouse: model.Warehouse = await service.get_warehouse_service(db, {'id': product.warehouse_id})
    new_products = [p.model_dump() for p in warehouse.products if p.id != product.id]
    await service.update_warehouse_service(db, warehouse_id=warehouse.id, warehouse_update={'products': new_products})


async def upload_excel_products_controller(db: DataBaseDep, products: UploadFile) -> None:
    [_, extension] = os.path.splitext(products.filename)
    if extension[1:] not in ['xlsx', 'xlsm']:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                'The files with extension ',
                f'"{extension[1:]}" are not supported.'
            )
        )
    fields_excel = ['nombre', 'cantidad', 'fecha caducidad', 'almacen']
    wb = openpyxl.load_workbook(products.file)
    ws = wb.active
    first_row = [
        ws.cell(row=1, column=i).value
        for i in range(1, len(fields_excel) + 1)
    ]
    if len(first_row) != len(fields_excel) and not all(
            field in fields_excel for field in first_row):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail='The excel file is incorrect'
        )
    products_excel: dict[str, model.WarehouseUpdate] = {}
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=4, values_only=True):
        if all(value is None for value in row):
            continue
        if row[0] is None or row[1] is None or row[3] is None:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail='The excel file is incorrect'
            )
        warehouse_name: str = str(row[3])
        try:
            new_product = model.Product(
                id=uuid4(),
                name=row[0],
                quantity=row[1],
                exp_date=row[2],
            )
        except ValidationError as e:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=parse_validation_error(e.errors())
            )
        if warehouse_name not in products_excel:
            products_excel[warehouse_name] = model.WarehouseUpdate(products=[])
        if new_product.name in [
            p.name for p in products_excel.get(warehouse_name).products
        ]:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail='There cannot be duplicated products'
            )
        products_excel.get(warehouse_name).products.append(new_product)
    warehouse_operations: list[BulkOperation] = []
    for key, value in products_excel.items():
        warehouse = await service.get_warehouse_service(db, query={'name': key})
        if warehouse is None:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f'Warehouse {key} not found'
            )
        products_names = [p.name for p in value.products]
        update_and_old_products: list[dict] = []
        for product in warehouse.products:
            if product.name in products_names:
                new_p = next(
                    (p for p in value.products if p.name == product.name),
                    None
                )
                new_p.id = product.id
                update_and_old_products.append(
                    get_valid_mongo_obj(new_p.model_dump()))
                value.products.remove(new_p)
                continue
            update_and_old_products.append(
                get_valid_mongo_obj(product.model_dump())
            )
        new_products = update_and_old_products + [
            get_valid_mongo_obj(p.model_dump()) for p in value.products
        ]
        warehouse_operations.append(
            BulkOperation(
                bulk_type='UpdateOne',
                data={'$set': {'products': new_products}},
                query=model.Warehouse.prepare_query({'id': warehouse.id})
            )
        )
    if len(warehouse_operations) > 0:
        await service.bulk_service(
            db,
            operations=warehouse_operations,
            ordered=False
        )
