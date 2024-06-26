from pathlib import Path
from uuid import uuid4

import openpyxl
import pytest_asyncio
from fastapi.testclient import TestClient
from pymongo.database import Database
from httpx import Response

from src.core.config import settings

URL_WAREHOUSE = f'{settings.API_STR}cyc/warehouse'


@pytest_asyncio.fixture
async def insert_warehouses_with_products(mongo_db: Database):
    mongo_db['Warehouse'].delete_many({})

    warehouse_id_1 = uuid4()
    warehouse_id_2 = uuid4()
    warehouse_id_3 = uuid4()
    warehouses = [
        {"_id": warehouse_id_1,
            "name": "Almacén 1",
            "products": [
                {
                    "id": uuid4(),
                    "name": "Producto 1",
                    "quantity": 10,
                    "exp_date": "2025-01-01",
                }
            ]
         },
        {"_id": warehouse_id_2,
            "name": "Almacén 2",
            "products": [
                {
                    "id": uuid4(),
                    "name": "Producto 2",
                    "quantity": 10,
                    "exp_date": "2025-01-01",
                }
            ]
         },
        {"_id": warehouse_id_3,
            "name": "Sevilla Este",
            "products": []
         }
    ]
    for warehouse in warehouses:
        mongo_db['Warehouse'].insert_one(warehouse)

    yield warehouses


def test_get_all_products_list(
    app_client: TestClient,
    insert_warehouses_with_products,
    app_superuser
):
    access_token = app_superuser['access_token']
    headers = {'authorization': f'Bearer {access_token}'}
    warehouses = insert_warehouses_with_products
    url = f'{URL_WAREHOUSE}/product'
    response: Response = app_client.get(url=url, headers=headers)
    assert response.status_code == 200
    result = response.json()
    assert 'elements' in result and 'total_elements' in result
    inserted_products = [
        product for warehouse in warehouses for product in warehouse["products"]]
    assert len(result['elements']) == len(inserted_products)
    for product_data in result['elements']:
        assert any(product_data["name"] == product["name"] and product_data["quantity"]
                   == product["quantity"] for product in inserted_products)


def test_get_product(
    app_client: TestClient,
    insert_warehouses_with_products,
    app_superuser
):
    access_token = app_superuser['access_token']
    headers = {'authorization': f'Bearer {access_token}'}
    product = insert_warehouses_with_products[0]['products'][0]
    url = f"{URL_WAREHOUSE}/product/{product['id']}"
    response: Response = app_client.get(url=url, headers=headers)
    assert response.status_code == 200
    result = response.json()
    assert result['id'] == str(product['id'])
    assert result['name'] == product['name']
    assert result['quantity'] == product['quantity']
    assert result['exp_date'] == product['exp_date']
    assert 'warehouse_id' in result


def test_create_product(
    app_client: TestClient,
    insert_warehouses_with_products,
    app_superuser
):
    access_token = app_superuser['access_token']
    headers = {'authorization': f'Bearer {access_token}'}
    warehouse_id = str(insert_warehouses_with_products[0]["_id"])
    url = f'{URL_WAREHOUSE}/product/'
    product_data = {
        "products": [
            {
                "name": "Queso",
                "exp_date": "2026-03-16",
                "quantity": 23,
                "warehouse_id": warehouse_id
            }
        ]
    }
    response = app_client.post(url=url, json=product_data, headers=headers)
    assert response.status_code == 201
    result = response.json()
    for field in product_data["products"][0]:
        assert str(result[0][field]) == str(product_data["products"][0][field])


def test_create_warehouse(app_client: TestClient, app_superuser):
    access_token = app_superuser['access_token']
    headers = {'authorization': f'Bearer {access_token}'}
    warehouse_data = {
        "name": "Almacén secundario",
        "products": [
            {
                "name": "Leche entera",
                "quantity": 34,
                "exp_date": "2025-03-16"
            }
        ]
    }
    response = app_client.post(
        url=URL_WAREHOUSE, json=warehouse_data, headers=headers)
    assert response.status_code == 201
    result = response.json()
    assert str(result['name']) == str(warehouse_data['name'])


def test_upload_excel_products(
        app_client: TestClient,
        mongo_db: Database,
        app_superuser):
    access_token = app_superuser['access_token']
    headers = {'authorization': f'Bearer {access_token}'}
    url = f'{URL_WAREHOUSE}/product/excel'
    # Load excel file
    excel_file_path = Path(__file__).resolve(
    ).parent.parent / 'excel_test' / 'Almacenes.xlsx'
    # Send excel file to endpoint
    with open(excel_file_path, 'rb') as file:
        files = {
            "products": (
                "Almacenes.xlsx",
                file,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        response: Response = app_client.post(
            url=url, files=files, headers=headers)
    # Verify response
    assert response.status_code == 204
    # Load excel file in memory
    wb = openpyxl.load_workbook(excel_file_path)
    ws = wb.active
    # Get db data
    warehouse_db = mongo_db["Warehouse"].find_one(
        {"name": ws.cell(row=2, column=4).value})
    products_db = list(warehouse_db["products"])
    first_product = products_db[0]
    # Compare db data with excel data
    assert first_product['name'] == ws.cell(row=2, column=1).value
    assert first_product['quantity'] == ws.cell(row=2, column=2).value
    assert first_product['exp_date'] == ws.cell(
        row=2, column=3
    ).value.date().isoformat()


def test_update_product(
    app_client: TestClient,
    insert_warehouses_with_products,
    app_superuser
):
    access_token = app_superuser['access_token']
    headers = {'authorization': f'Bearer {access_token}'}
    warehouse_id = str(insert_warehouses_with_products[0]["_id"])
    product_id = str(insert_warehouses_with_products[0]["products"][0]["id"])
    url = f'{URL_WAREHOUSE}/product/'
    product_data = {
        "products": [
            {
                "name": "Queso",
                "exp_date": "2027-03-16",
                "quantity": 23,
                "update_exp_date": 'false',
                "warehouse_id": str(
                    insert_warehouses_with_products[0]["_id"]),
                "product_id": str(
                    insert_warehouses_with_products[0]["products"][0]["id"]),
            }]}
    response: Response = app_client.patch(
        url=url, json=product_data, headers=headers
    )
    assert response.status_code == 200
    result = response.json()
    assert str(result[0]["exp_date"]) == "2025-01-01"
    assert str(result[0]["name"]) == "Queso"
    assert str(result[0]["quantity"]) == "23"
    assert str(result[0]["warehouse_id"]) == warehouse_id
    assert str(result[0]["id"]) == product_id


def test_update_warehouse(
    app_client: TestClient,
    insert_warehouses_with_products,
    app_superuser
):
    access_token = app_superuser['access_token']
    headers = {'authorization': f'Bearer {access_token}'}
    warehouse_id = str(insert_warehouses_with_products[1]["_id"])
    url = f'{URL_WAREHOUSE}/{warehouse_id}/'
    warehouse_data = {
        "name": "Almacen",
        "products": []
    }
    response: Response = app_client.patch(
        url=url, json=warehouse_data, headers=headers
    )
    assert response.status_code == 200
    result = response.json()
    assert str(result["name"]) == "Almacen"
    assert len(result['products']) == 0


def test_delete_product(
    app_client: TestClient,
    insert_warehouses_with_products,
    app_superuser
):
    access_token = app_superuser['access_token']
    headers = {'authorization': f'Bearer {access_token}'}
    product_id = insert_warehouses_with_products[0]["products"][0]["id"]
    url = f'{URL_WAREHOUSE}/product/{product_id}'
    response = app_client.delete(url=url, headers=headers)
    assert response.status_code == 204
    response = app_client.get(url=url, headers=headers)
    assert response.status_code == 404
